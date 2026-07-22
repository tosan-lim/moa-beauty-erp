import sqlite3
import pandas as pd
from datetime import datetime
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st

# ==========================================
# 1. 페이지 및 다국어 설정
# ==========================================
st.set_page_config(
    page_title="KIL INDIA - Moa Beauty ERP",
    page_icon="🧴",
    layout="wide"
)

LANG = {
    'ko': {
        'title': "KIL INDIA TRADE PVT. LTD. - Moa Beauty ERP",
        'sub': "인도 법인 화장품 재고 및 GST/관세 통합 관리 웹 시스템",
        'menu_master': "📋 기초등록 (Master)",
        'menu_purchase': "📦 구매/입고 (Purchase)",
        'menu_sales': "🏷️ 판매/출고 (Sales)",
        'menu_reports': "📊 엑셀 보고서 (Reports)",
        
        # Master
        'p_reg': "거래처 상세 등록",
        'p_code': "거래처코드", 'p_name': "상호명", 'p_type': "구분",
        'p_phone': "전화번호", 'p_email': "이메일", 'p_state': "State(주)",
        'p_addr': "상세주소", 'p_contact': "담당자명", 'p_cat': "카테고리",
        'btn_save_p': "거래처 저장",
        'loc_reg': "보관처(창고) 등록", 'item_reg': "품목 등록",

        # Purchase
        'pur_title': "구매 및 수입 입고 (Customs Duty & GST)",
        'pur_type': "입고 구분",
        'duty_rate': "관세율 (Customs Duty %)",
        'gst_rate': "GST 세율 (%)",
        'qty': "수량", 'price': "단가 (INR ₹)", 'lot': "LOT 번호", 'exp': "유통기한 (YYYY-MM-DD)",
        'handler': "담당자", 'remarks': "비고(특이사항)",
        'btn_save_pur': "입고 등록 실행",

        # Sales
        'sal_title': "판매 및 출고 (FEFO 유통기한 우선 차감)",
        'gst_choice': "GST 적용 여부",
        'btn_save_sal': "출고 등록 실행",

        # Reports
        'rep_title': "엑셀 리포트 다운로드",
        'btn_inv': "📊 현재고 보고서 다운로드",
        's_date': "시작일자", 'e_date': "종료일자",
        'btn_sales_rep': "📑 기간별 매출 명세서 (Sales GST) 다운로드",
        'btn_pur_rep': "📦 기간별 매입/수입 명세서 (Customs & GST) 다운로드"
    },
    'en': {
        'title': "KIL INDIA TRADE PVT. LTD. - Moa Beauty ERP",
        'sub': "India Branch Cosmetics Inventory & GST/Customs ERP Web System",
        'menu_master': "📋 Master Data",
        'menu_purchase': "📦 Purchase / Import",
        'menu_sales': "🏷️ Sales / Stock-Out",
        'menu_reports': "📊 Excel Reports",

        # Master
        'p_reg': "Partner Registration",
        'p_code': "Partner Code", 'p_name': "Company Name", 'p_type': "Type",
        'p_phone': "Phone", 'p_email': "Email", 'p_state': "State",
        'p_addr': "Address Detail", 'p_contact': "Contact Person", 'p_cat': "Category",
        'btn_save_p': "Save Partner",
        'loc_reg': "Warehouse Registration", 'item_reg': "Item Registration",

        # Purchase
        'pur_title': "Purchase & Import Stock-In (Duty & GST)",
        'pur_type': "Purchase Type",
        'duty_rate': "Customs Duty Rate (%)",
        'gst_rate': "GST Rate (%)",
        'qty': "Quantity", 'price': "Unit Price (INR ₹)", 'lot': "LOT Number", 'exp': "Expiry Date (YYYY-MM-DD)",
        'handler': "Handler", 'remarks': "Remarks",
        'btn_save_pur': "Execute Stock-In",

        # Sales
        'sal_title': "Sales & Stock-Out (FEFO Auto Deduction)",
        'gst_choice': "GST Applicable",
        'btn_save_sal': "Execute Sales",

        # Reports
        'rep_title': "Export Excel Reports",
        'btn_inv': "📊 Download Current Inventory Report",
        's_date': "Start Date", 'e_date': "End Date",
        'btn_sales_rep': "📑 Download Sales GST Report",
        'btn_pur_rep': "📦 Download Purchase & Customs Report"
    }
}

# ==========================================
# 2. 데이터베이스 설정 및 자동 보정
# ==========================================
DB_NAME = "kil_india_erp.db"

def get_connection():
    return sqlite3.connect(DB_NAME)

def init_db():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS partners (
        partner_code TEXT PRIMARY KEY, partner_name TEXT NOT NULL, partner_type TEXT NOT NULL,
        phone TEXT, email TEXT, state TEXT, address_detail TEXT, contact_person TEXT, category TEXT
    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS locations (
        location_code TEXT PRIMARY KEY, location_name TEXT NOT NULL
    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS items (
        item_code TEXT PRIMARY KEY, item_name TEXT NOT NULL, base_price REAL, safety_stock INTEGER DEFAULT 0
    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS inventory (
        inv_id INTEGER PRIMARY KEY AUTOINCREMENT, location_code TEXT, item_code TEXT, lot_number TEXT, expiry_date TEXT, quantity INTEGER
    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS purchases (
        purchase_id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, partner_code TEXT, location_code TEXT, item_code TEXT,
        quantity INTEGER, unit_price REAL, subtotal REAL, duty_rate REAL, duty_amount REAL,
        gst_type TEXT, gst_rate REAL, gst_amount REAL, total_amount REAL, handler TEXT, remarks TEXT
    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS sales (
        sale_id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, partner_code TEXT, location_code TEXT, item_code TEXT,
        quantity INTEGER, unit_price REAL, subtotal REAL, gst_type TEXT, gst_rate REAL, gst_amount REAL, total_amount REAL, handler TEXT, remarks TEXT
    )''')
    conn.commit()

    def add_col(tbl, col):
        try: cursor.execute(f"ALTER TABLE {tbl} ADD COLUMN {col}")
        except: pass
    for c in ["phone TEXT", "email TEXT", "state TEXT", "address_detail TEXT", "contact_person TEXT", "category TEXT"]: add_col("partners", c)
    for c in ["subtotal REAL", "duty_rate REAL", "duty_amount REAL", "gst_type TEXT", "gst_rate REAL", "gst_amount REAL", "handler TEXT", "remarks TEXT"]: add_col("purchases", c)
    for c in ["subtotal REAL", "gst_type TEXT", "gst_rate REAL", "gst_amount REAL", "handler TEXT", "remarks TEXT"]: add_col("sales", c)
    conn.commit()
    conn.close()

init_db()

# ==========================================
# 3. 사이드바 다국어 지원
# ==========================================
st.sidebar.image("https://img.icons8.com/color/96/000000/top-cream.png", width=70)
st.sidebar.title("KIL INDIA TRADE")
lang_choice = st.sidebar.radio("🌐 Language / 언어 선택", ["한국어 🇰🇷", "English 🇬🇧"])
lang_code = 'ko' if "한국어" in lang_choice else 'en'
L = LANG[lang_code]

st.title(L['title'])
st.caption(L['sub'])
st.divider()

# ==========================================
# 4. 웹 메인 탭 메뉴
# ==========================================
tab1, tab2, tab3, tab4 = st.tabs([L['menu_master'], L['menu_purchase'], L['menu_sales'], L['menu_reports']])

# --- TAB 1: 기초등록 ---
with tab1:
    st.subheader(L['p_reg'])
    col1, col2, col3 = st.columns(3)
    with col1:
        p_code = st.text_input(L['p_code'], key="p_code")
        p_phone = st.text_input(L['p_phone'], key="p_phone")
        p_addr = st.text_input(L['p_addr'], key="p_addr")
    with col2:
        p_name = st.text_input(L['p_name'], key="p_name")
        p_email = st.text_input(L['p_email'], key="p_email")
        p_contact = st.text_input(L['p_contact'], key="p_contact")
    with col3:
        p_type = st.selectbox(L['p_type'], ["매출처/Buyer", "매입처/Vendor"])
        p_state = st.text_input(L['p_state'], key="p_state")
        p_cat = st.text_input(L['p_cat'], key="p_cat")

    if st.button(L['btn_save_p'], type="primary"):
        if p_code and p_name:
            conn = get_connection()
            cur = conn.cursor()
            try:
                cur.execute("INSERT INTO partners VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                            (p_code, p_name, p_type, p_phone, p_email, p_state, p_addr, p_contact, p_cat))
                conn.commit()
                st.success(f"✅ Partner [{p_name}] Registered Successfully!")
            except Exception as e: st.error(f"Error: {e}")
            finally: conn.close()
        else:
            st.warning("Please fill Code & Name.")

    st.divider()
    subcol1, subcol2 = st.columns(2)
    with subcol1:
        st.subheader(L['loc_reg'])
        l_code = st.text_input("Warehouse Code / 보관처코드")
        l_name = st.text_input("Warehouse Name / 보관처명")
        if st.button("Save Warehouse / 창고 저장"):
            if l_code and l_name:
                conn = get_connection()
                cur = conn.cursor()
                cur.execute("INSERT INTO locations VALUES (?, ?)", (l_code, l_name))
                conn.commit()
                conn.close()
                st.success("Saved Warehouse!")

    with subcol2:
        st.subheader(L['item_reg'])
        i_code = st.text_input("Item Code / 품목코드")
        i_name = st.text_input("Item Name / 품목명")
        i_price = st.number_input("Base Price / 기준단가 (INR ₹)", min_value=0.0, value=0.0)
        if st.button("Save Item / 품목 저장"):
            if i_code and i_name:
                conn = get_connection()
                cur = conn.cursor()
                cur.execute("INSERT INTO items VALUES (?, ?, ?, 0)", (i_code, i_name, i_price))
                conn.commit()
                conn.close()
                st.success("Saved Item!")

# --- TAB 2: 구매/입고 ---
with tab2:
    st.subheader(L['pur_title'])
    p_col1, p_col2, p_col3 = st.columns(3)
    with p_col1:
        pur_p_code = st.text_input(L['p_code'], key="pur_p_code")
        pur_item_code = st.text_input("Item Code / 품목코드", key="pur_item_code")
        pur_qty = st.number_input(L['qty'], min_value=1, value=100, key="pur_qty")
        pur_duty = st.number_input(L['duty_rate'], min_value=0.0, value=10.0)
    with p_col2:
        pur_loc_code = st.text_input("Location Code / 보관처코드", key="pur_loc_code")
        pur_lot = st.text_input(L['lot'], key="pur_lot")
        pur_price = st.number_input(L['price'], min_value=0.0, value=250.0, key="pur_price")
        pur_gst = st.number_input(L['gst_rate'], min_value=0.0, value=18.0, key="pur_gst")
    with p_col3:
        pur_type = st.selectbox(L['pur_type'], ["해외 수입 (Import)", "인도 국내 (Domestic)", "면세 (NO Tax)"])
        pur_exp = st.text_input(L['exp'], value="2028-12-31")
        pur_handler = st.text_input(L['handler'], key="pur_handler")
        pur_remarks = st.text_input(L['remarks'], key="pur_remarks")

    if st.button(L['btn_save_pur'], type="primary"):
        subtotal = round(pur_qty * pur_price, 2)
        duty_amt = round(subtotal * (pur_duty / 100.0), 2) if "Import" in pur_type or "수입" in pur_type else 0.0
        gst_amt = round((subtotal + duty_amt) * (pur_gst / 100.0), 2) if "Tax" not in pur_type and "면세" not in pur_type else 0.0
        total_amt = round(subtotal + duty_amt + gst_amt, 2)
        today = datetime.now().strftime('%Y-%m-%d')

        conn = get_connection()
        cur = conn.cursor()
        cur.execute('''INSERT INTO purchases (date, partner_code, location_code, item_code, quantity, unit_price, subtotal, duty_rate, duty_amount, gst_type, gst_rate, gst_amount, total_amount, handler, remarks)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (today, pur_p_code, pur_loc_code, pur_item_code, pur_qty, pur_price, subtotal, pur_duty, duty_amt, pur_type, pur_gst, gst_amt, total_amt, pur_handler, pur_remarks))
        cur.execute("INSERT INTO inventory (location_code, item_code, lot_number, expiry_date, quantity) VALUES (?, ?, ?, ?, ?)",
                    (pur_loc_code, pur_item_code, pur_lot, pur_exp, pur_qty))
        conn.commit()
        conn.close()
        st.success(f"✅ Stock-In Complete! Subtotal: ₹{subtotal:,.2f} | Duty: ₹{duty_amt:,.2f} | GST: ₹{gst_amt:,.2f} | Total: ₹{total_amt:,.2f}")

# --- TAB 3: 판매/출고 ---
with tab3:
    st.subheader(L['sal_title'])
    s_col1, s_col2, s_col3 = st.columns(3)
    with s_col1:
        sal_p_code = st.text_input(L['p_code'], key="sal_p_code")
        sal_qty = st.number_input(L['qty'], min_value=1, value=10, key="sal_qty")
        sal_gst_choice = st.selectbox(L['gst_choice'], ["GST", "NO GST"])
    with s_col2:
        sal_loc_code = st.text_input("Location Code / 보관처코드", key="sal_loc_code")
        sal_price = st.number_input(L['price'], min_value=0.0, value=350.0, key="sal_price")
        sal_gst_rate = st.number_input(L['gst_rate'], min_value=0.0, value=18.0, key="sal_gst_rate")
    with s_col3:
        sal_item_code = st.text_input("Item Code / 품목코드", key="sal_item_code")
        sal_handler = st.text_input(L['handler'], key="sal_handler")
        sal_remarks = st.text_input(L['remarks'], key="sal_remarks")

    if st.button(L['btn_save_sal'], type="primary"):
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT inv_id, lot_number, expiry_date, quantity FROM inventory WHERE location_code=? AND item_code=? AND quantity>0 ORDER BY expiry_date ASC", (sal_loc_code, sal_item_code))
        lots = cur.fetchall()
        total_avail = sum(l[3] for l in lots)

        if total_avail < sal_qty:
            st.error(f"❌ Stock Error! Current Stock: {total_avail} / Requested: {sal_qty}")
            conn.close()
        else:
            rem = sal_qty
            for inv_id, lot_num, exp, lot_qty in lots:
                if rem <= 0: break
                deduct = min(lot_qty, rem)
                cur.execute("UPDATE inventory SET quantity = quantity - ? WHERE inv_id = ?", (deduct, inv_id))
                rem -= deduct

            subtotal = round(sal_qty * sal_price, 2)
            gst_rate = sal_gst_rate if sal_gst_choice == "GST" else 0.0
            gst_amt = round(subtotal * (gst_rate / 100.0), 2)
            total_amt = round(subtotal + gst_amt, 2)
            today = datetime.now().strftime('%Y-%m-%d')

            cur.execute('''INSERT INTO sales (date, partner_code, location_code, item_code, quantity, unit_price, subtotal, gst_type, gst_rate, gst_amount, total_amount, handler, remarks)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        (today, sal_p_code, sal_loc_code, sal_item_code, sal_qty, sal_price, subtotal, sal_gst_choice, gst_rate, gst_amt, total_amt, sal_handler, sal_remarks))
            conn.commit()
            conn.close()
            st.success(f"✅ FEFO Stock-Out Complete! Subtotal: ₹{subtotal:,.2f} | GST: ₹{gst_amt:,.2f} | Total: ₹{total_amt:,.2f}")

# --- TAB 4: 엑셀 보고서 ---
with tab4:
    st.subheader(L['rep_title'])

    def generate_excel_bytes(df, sheet_title, main_title):
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        col_span = len(df.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
        title_cell = ws.cell(row=1, column=1, value=main_title)
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="1F4E78")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for c_idx, col_name in enumerate(df.columns, 1):
            c = ws.cell(row=3, column=c_idx, value=col_name)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max(len(str(col_name)) * 2, 14)

        for r_idx, row in enumerate(df.values, 4):
            for c_idx, val in enumerate(row, 1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.border = border
                if isinstance(val, (int, float)):
                    if "수량" in df.columns[c_idx-1]:
                        c.number_format = '#,##0'
                    else:
                        c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(output)
        return output.getvalue()

    # 1. 현재고 보고서
    conn = get_connection()
    q_inv = '''SELECT loc.location_name as '보관처(창고)', itm.item_code as '품목코드', itm.item_name as '품목명',
                      COALESCE(SUM(inv.quantity), 0) as '현재고(개)', itm.safety_stock as '안전재고(개)'
               FROM items itm CROSS JOIN locations loc
               LEFT JOIN inventory inv ON itm.item_code = inv.item_code AND loc.location_code = inv.location_code
               GROUP BY loc.location_code, itm.item_code'''
    df_inv = pd.read_sql(q_inv, conn)
    df_inv = df_inv[df_inv['현재고(개)'] > 0]
    conn.close()

    st.write("#### 1. Current Inventory Status")
    st.dataframe(df_inv, use_container_width=True)
    if not df_inv.empty:
        st.download_button(L['btn_inv'], data=generate_excel_bytes(df_inv, "Current Inventory", "KIL INDIA - Current Inventory Report"), file_name="Moa_Beauty_Current_Inventory.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.divider()

    # 2. 기간별 리포트
    st.write("#### 2. Periodical Tax & Customs Reports")
    d_col1, d_col2 = st.columns(2)
    with d_col1: s_date = st.text_input(L['s_date'], "2026-01-01")
    with d_col2: e_date = st.text_input(L['e_date'], "2026-12-31")

    # Sales Report
    conn = get_connection()
    q_sales = f'''SELECT s.date as '일자', s.partner_code as '거래처코드', COALESCE(p.partner_name, s.partner_code) as '거래처명',
                        COALESCE(p.state, '-') as 'State(주)', s.location_code as '출고창고', s.item_code as '품목코드',
                        COALESCE(i.item_name, s.item_code) as '품목명', s.quantity as '수량', s.unit_price as '단가(INR ₹)',
                        s.subtotal as '공급가액(INR ₹)', s.gst_type as 'GST구분', s.gst_rate as 'GST세율(%)', s.gst_amount as 'GST세액(INR ₹)',
                        s.total_amount as '총합계(INR ₹)', s.handler as '담당자', s.remarks as '비고'
                 FROM sales s LEFT JOIN partners p ON s.partner_code = p.partner_code LEFT JOIN items i ON s.item_code = i.item_code
                 WHERE s.date BETWEEN '{s_date}' AND '{e_date}' ORDER BY s.date ASC'''
    df_sales = pd.read_sql(q_sales, conn)

    # Purchase Report
    q_pur = f'''SELECT pur.date as '일자', pur.partner_code as '매입처코드', COALESCE(p.partner_name, pur.partner_code) as '매입처명',
                       COALESCE(p.state, '-') as 'State(주)', pur.location_code as '입고창고', pur.item_code as '품목코드',
                       COALESCE(i.item_name, pur.item_code) as '품목명', pur.quantity as '수량', pur.unit_price as '단가(INR ₹)',
                       COALESCE(pur.subtotal, pur.total_amount) as '공급가액(INR ₹)', COALESCE(pur.duty_rate, 0) as '관세율(%)',
                       COALESCE(pur.duty_amount, 0) as '관세액(INR ₹)', COALESCE(pur.gst_type, 'NO GST') as 'GST구분',
                       COALESCE(pur.gst_rate, 0) as 'GST세율(%)', COALESCE(pur.gst_amount, 0) as 'GST세액(INR ₹)',
                       pur.total_amount as '총매입액(INR ₹)', pur.handler as '담당자', pur.remarks as '비고'
                FROM purchases pur LEFT JOIN partners p ON pur.partner_code = p.partner_code LEFT JOIN items i ON pur.item_code = i.item_code
                WHERE pur.date BETWEEN '{s_date}' AND '{e_date}' ORDER BY pur.date ASC'''
    df_pur = pd.read_sql(q_pur, conn)
    conn.close()

    r_col1, r_col2 = st.columns(2)
    with r_col1:
        if not df_sales.empty:
            st.download_button(L['btn_sales_rep'], data=generate_excel_bytes(df_sales, "Sales GST Summary", f"KIL INDIA - GST Sales Report ({s_date} ~ {e_date})"), file_name=f"Moa_Beauty_Sales_GST_{s_date}_to_{e_date}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else: st.info("No sales data in period.")

    with r_col2:
        if not df_pur.empty:
            st.download_button(L['btn_pur_rep'], data=generate_excel_bytes(df_pur, "Purchase Duty Summary", f"KIL INDIA - Purchase & Customs Duty Report ({s_date} ~ {e_date})"), file_name=f"Moa_Beauty_Purchase_Duty_{s_date}_to_{e_date}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else: st.info("No purchase data in period.")